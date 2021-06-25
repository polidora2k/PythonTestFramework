import openpyxl
from multipledispatch import dispatch

"""
    HomePageData provides methods to get testing data for various
    test cases and form submission test.
"""
class HomePageData:

    """
        getData accesses an excel sheet containing test data to use for
        testing of the Home Page.  The function is overloaded with one
        version taking no arguments and the other taking one.  A specific
        data set can be chosen by supplying text or all test cases can be
        run.
    """
    @dispatch()
    def getData(self):
        sheet = openpyxl.load_workbook("/Users/ipolidora/PycharmProjects/PythonTestFramework/testData/testData.xlsx")["test_formSubmission"]
        testData = []

        for i in range(2, sheet.max_row + 1):
            testcaseData = {}
            for j in range(1, sheet.max_column + 1):
                testcaseData[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

            testData.append(testcaseData)

        return testData

    @dispatch(str)
    def getData(self, testName):
        sheet = openpyxl.load_workbook("/Users/ipolidora/PycharmProjects/PythonTestFramework/testData/testData.xlsx").get_sheet_by_name("test_formSubmission")
        testData = []

        for i in range(2, sheet.max_row + 1):
            testcaseData = {}
            if sheet.cell(row=i,column=1) == testName:
                for j in range(1, sheet.max_column + 1):
                    testcaseData[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value

                testData.append(testcaseData)

        return testData
